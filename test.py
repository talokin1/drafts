import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule


def create_golden_report(
    model,
    X_train,
    valid_pool,
    inference_result,
    metrics,
    cm,
    BEST_THRESHOLD,
    output_path="Golden_Model_Report.xlsx"
):
    wb = Workbook.create()

    # =========================
    # COLORS
    # =========================
    NAVY = "#1F4E78"
    BLUE = "#5B9BD5"
    LIGHT_BLUE = "#DDEBF7"
    LIGHTER_BLUE = "#EAF2F8"
    GREEN = "#E2F0D9"
    RED = "#F4CCCC"
    ORANGE = "#FCE4D6"
    WHITE = "#FFFFFF"

    def section(ws, rng, text, fill=NAVY, font_color=WHITE, size=12):
        ws.merge_cells(rng)
        ws.get_range(rng.split(":")[0]).values = [[text]]
        ws.get_range(rng).format = {
            "fill": fill,
            "font": {"bold": True, "color": font_color, "size": size},
            "horizontal_alignment": "center",
            "vertical_alignment": "center",
            "wrap_text": True
        }

    # ============================================================
    # 1. MODEL DESCRIPTION
    # ============================================================
    ws = wb.worksheets.add("Model description")

    section(ws, "B2:H2", "Task description")

    ws.merge_cells("B3:H6")
    ws.get_range("B3").values = [[
        "The goal of the model is to estimate how similar an HNWI client's "
        "current profile is to Golden clients. The model is used to rank "
        "currently non-Golden HNWI clients by their Golden propensity."
    ]]
    ws.get_range("B3:H6").format = {
        "fill": LIGHTER_BLUE,
        "wrap_text": True,
        "vertical_alignment": "center"
    }

    section(ws, "B8:H8", "Target description")

    ws.merge_cells("B9:H11")
    ws.get_range("B9").values = [[
        "TARGET = 1 (Golden): PACKAGE criterion is mandatory and at least "
        "2 additional Golden criteria must be satisfied. "
        "TARGET = 0: otherwise."
    ]]
    ws.get_range("B9:H11").format = {"fill": LIGHTER_BLUE, "wrap_text": True}

    ws.get_range("B13:C18").values = [
        ["Criterion", "Condition"],
        ["PACKAGE", "Mandatory"],
        ["TOTAL_PORTFOLIO", "> 4,000,000"],
        ["LIABILITIES_UAH", "> 1,000,000"],
        ["INCOME(COM+INTEREST)", "> 15,000"],
        ["AMT_DEB_CARD", "> 50,000"]
    ]

    ws.get_range("B13:C13").format = {
        "fill": NAVY,
        "font": {"bold": True, "color": WHITE},
        "horizontal_alignment": "center"
    }

    section(ws, "B20:H20", "Model output")

    ws.get_range("B22:C24").values = [
        ["GOLDEN_SCORE", "Continuous Golden propensity score from 0 to 1"],
        ["MODEL_CLASS", "Golden / Not Golden based on selected threshold"],
        ["GOLDEN_RANK", "Client rank by Golden propensity"]
    ]

    ws.get_range("B2:H24").format.wrap_text = True
    ws.get_range("B:B").format.column_width = 27
    ws.get_range("C:H").format.column_width = 18

    # ============================================================
    # 2. ARCHITECTURE
    # ============================================================
    ws = wb.worksheets.add("Architecture")

    section(ws, "B2:G2", "Architecture CatBoost")

    ws.merge_cells("B4:G7")
    ws.get_range("B4").values = [[
        "The model receives current HNWI client characteristics and estimates "
        "a continuous Golden propensity score. The score is then converted "
        "into Golden / Not Golden using the selected classification threshold."
    ]]
    ws.get_range("B4:G7").format = {
        "fill": LIGHTER_BLUE,
        "wrap_text": True,
        "horizontal_alignment": "center",
        "vertical_alignment": "center"
    }

    section(ws, "B9:G10", "INPUT DATA\nHNWI client × features")
    section(ws, "B12:G13", "CatBoostClassifier\nBinary classification", BLUE)
    section(ws, "B15:G16", "GOLDEN_SCORE\nProbability / propensity from 0 to 1")
    section(
        ws,
        "B18:G19",
        f"CLASSIFICATION THRESHOLD\n{BEST_THRESHOLD:.2%}",
        ORANGE,
        NAVY
    )
    section(ws, "B21:D22", "NOT GOLDEN", RED, NAVY)
    section(ws, "E21:G22", "GOLDEN", GREEN, NAVY)

    for row in [11, 14, 17, 20]:
        ws.merge_cells(f"B{row}:G{row}")
        ws.get_range(f"B{row}").values = [["↓"]]
        ws.get_range(f"B{row}:G{row}").format = {
            "font": {"bold": True, "size": 18, "color": NAVY},
            "horizontal_alignment": "center"
        }

    ws.get_range("B:G").format.column_width = 13

    # ============================================================
    # 3. EXAMPLE
    # ============================================================
    ws = wb.worksheets.add("Example")

    example = (
        inference_result[
            ["CONTRAGENTID", "GOLDEN_SCORE", "MODEL_CLASS", "GOLDEN_RANK"]
        ]
        .head(10)
        .copy()
    )

    headers = ["CLIENT_ID", "Golden propensity", "Model class", "Rank"]

    ws.get_range("B2:E2").values = [headers]
    ws.get_range("B2:E2").format = {
        "fill": NAVY,
        "font": {"bold": True, "color": WHITE},
        "horizontal_alignment": "center"
    }

    ws.get_range(f"B3:E{len(example)+2}").values = example.values.tolist()
    ws.get_range(f"C3:C{len(example)+2}").format.number_format = "0.00%"

    ws.get_range("B:B").format.column_width = 16
    ws.get_range("C:C").format.column_width = 20
    ws.get_range("D:D").format.column_width = 18
    ws.get_range("E:E").format.column_width = 10

    # ============================================================
    # 4. MODEL RESULT
    # ============================================================
    ws = wb.worksheets.add("Model result")

    section(ws, "B2:H2", "Model result")

    ws.get_range("B4:C9").values = [
        ["Metric", "Value"],
        ["Best threshold", float(BEST_THRESHOLD)],
        ["Precision", float(metrics["Precision"])],
        ["Recall", float(metrics["Recall"])],
        ["F1", float(metrics["F1"])],
        ["ROC-AUC", float(metrics["ROC_AUC"])]
    ]

    ws.get_range("B4:C4").format = {
        "fill": NAVY,
        "font": {"bold": True, "color": WHITE},
        "horizontal_alignment": "center"
    }

    ws.get_range("C5:C9").format.number_format = "0.00%"

    section(ws, "E4:G4", "Confusion matrix")

    ws.get_range("E5:G7").values = [
        ["", "Pred Not Golden", "Pred Golden"],
        ["Actual Not Golden", int(cm.iloc[0, 0]), int(cm.iloc[0, 1])],
        ["Actual Golden", int(cm.iloc[1, 0]), int(cm.iloc[1, 1])]
    ]

    ws.get_range("E5:G5").format = {
        "fill": NAVY,
        "font": {"bold": True, "color": WHITE},
        "horizontal_alignment": "center"
    }

    ws.get_range("E6:E7").format = {
        "fill": LIGHT_BLUE,
        "font": {"bold": True}
    }

    ws.merge_cells("B12:H14")
    ws.get_range("B12").values = [[
        "Metrics are calculated on out-of-fold predictions. "
        "The classification threshold is selected by maximum F1 score."
    ]]
    ws.get_range("B12:H14").format = {
        "fill": LIGHTER_BLUE,
        "wrap_text": True,
        "vertical_alignment": "center"
    }

    ws.get_range("B:B").format.column_width = 22
    ws.get_range("C:C").format.column_width = 15
    ws.get_range("E:G").format.column_width = 20

    # ============================================================
    # 5. FEATURE IMPORTANCE
    # ============================================================
    ws = wb.worksheets.add("Feature importance")

    fi = model.get_feature_importance(
        valid_pool,
        type="PredictionValuesChange",
        prettified=True
    ).copy()

    total_importance = fi["Importances"].sum()
    fi["Importance_%"] = fi["Importances"] / total_importance

    values = [["Feature", "Importance", "Importance_%"]] + fi.values.tolist()

    ws.get_range(f"B2:D{len(values)+1}").values = values

    ws.get_range("B2:D2").format = {
        "fill": NAVY,
        "font": {"bold": True, "color": WHITE},
        "horizontal_alignment": "center"
    }

    ws.get_range(f"D3:D{len(values)+1}").format.number_format = "0.000%"

    ws.get_range(f"C3:C{len(values)+1}").conditional_formats.add_data_bar({
        "color": BLUE,
        "gradient": True
    })

    ws.get_range(f"D3:D{len(values)+1}").conditional_formats.add_data_bar({
        "color": "#70AD47",
        "gradient": True
    })

    ws.get_range("B:B").format.column_width = 38
    ws.get_range("C:D").format.column_width = 18
    ws.freeze_panes.freeze_rows(2)

    # ============================================================
    # 6. FEATURE DESCRIBE
    # ============================================================
    ws = wb.worksheets.add("Feature describe")

    description_rows = [["Feature", "Dtype", "Missing_%", "N_unique", "Description"]]

    for col in X_train.columns:
        description_rows.append([
            col,
            str(X_train[col].dtype),
            float(X_train[col].isna().mean()),
            int(X_train[col].nunique(dropna=False)),
            ""
        ])

    ws.get_range(f"B2:F{len(description_rows)+1}").values = description_rows

    ws.get_range("B2:F2").format = {
        "fill": NAVY,
        "font": {"bold": True, "color": WHITE},
        "horizontal_alignment": "center"
    }

    ws.get_range(f"D3:D{len(description_rows)+1}").format.number_format = "0.00%"

    ws.get_range("B:B").format.column_width = 38
    ws.get_range("C:E").format.column_width = 16
    ws.get_range("F:F").format.column_width = 40
    ws.freeze_panes.freeze_rows(2)

    # ============================================================
    # SAVE
    # ============================================================
    SpreadsheetFile.export_xlsx(wb).save(output_path)

    print(f"Report saved: {output_path}")



create_golden_report(
    model=model,
    X_train=X_train,
    valid_pool=valid_pool,
    inference_result=inference_result,
    metrics=metrics,
    cm=cm,
    BEST_THRESHOLD=BEST_THRESHOLD,
    output_path="Golden_Model_Report.xlsx"
)